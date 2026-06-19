"""
Calcula métricas completas de segmentación para resultados 2D y 3D de nnUNet.

Métricas calculadas por caso y de media:
  - Dice, IoU (del JSON)
  - Precision, Specificity (del JSON via TP/FP/TN/FN)
  - Volumen GT (mL), Volumen predicho (mL), Error Vol (%) (del JSON + espaciado NIfTI)
  - HD95, HDmax, SSIM (requiere cargar NIfTIs)

"""

import argparse
import json
import os
import warnings
from pathlib import Path
 
import numpy as np
import pandas as pd
 
warnings.filterwarnings("ignore")
 
 
# ------------------------------------------------------------------------------
# Helpers de métricas sobre NIfTI
# ------------------------------------------------------------------------------
 
def _get_surface(binary_mask: np.ndarray) -> np.ndarray:
    """Devuelve los vóxeles de superficie (erosión morfológica)."""
    from scipy.ndimage import binary_erosion
    if not np.any(binary_mask):
        return binary_mask
    eroded = binary_erosion(binary_mask, border_value=0)
    return binary_mask & ~eroded
 
 
def compute_hausdorff(pred_mask: np.ndarray,
                      ref_mask: np.ndarray,
                      spacing: tuple) -> tuple:
    """
    Calcula HD95 y HDmax simétrica entre dos máscaras binarias.
 
    Returns
    -------
    hd95 : float  (np.nan si alguna máscara está vacía)
    hdmax: float
    """
    from scipy.ndimage import distance_transform_edt
 
    if not np.any(pred_mask) or not np.any(ref_mask):
        return np.nan, np.nan
 
    pred_surf = _get_surface(pred_mask)
    ref_surf  = _get_surface(ref_mask)
 
    if not np.any(pred_surf) or not np.any(ref_surf):
        return np.nan, np.nan
 
    # Distancia euclidiana desde la superficie de cada máscara
    dt_from_pred = distance_transform_edt(~pred_surf, sampling=spacing)
    dt_from_ref  = distance_transform_edt(~ref_surf,  sampling=spacing)
 
    # Distancias simétricas: pred→ref y ref→pred
    d_pred_to_ref = dt_from_ref[pred_surf]
    d_ref_to_pred = dt_from_pred[ref_surf]
 
    all_d = np.concatenate([d_pred_to_ref, d_ref_to_pred])
    hd95  = float(np.percentile(all_d, 95))
    hdmax = float(np.max(all_d))
 
    return hd95, hdmax
 
 
def compute_ssim(pred_mask: np.ndarray, ref_mask: np.ndarray) -> float:
    """SSIM entre dos máscaras binarias (float32 en [0,1])."""
    from skimage.metrics import structural_similarity as skssim
 
    p = pred_mask.astype(np.float32)
    r = ref_mask.astype(np.float32)
 
    # Para volúmenes 3D usamos win_size pequeño si el patch es diminuto
    min_side = min(p.shape)
    win_size = min(7, min_side) if min_side >= 3 else 3
    if win_size % 2 == 0:
        win_size -= 1
 
    try:
        val = skssim(p, r, data_range=1.0, win_size=win_size)
    except Exception:
        val = np.nan
 
    return float(val)
 
 
# ------------------------------------------------------------------------------
# Métricas derivadas del JSON (sin cargar NIfTIs)
# ------------------------------------------------------------------------------
 
def metrics_from_counts(tp: float, fp: float, tn: float, fn: float,
                        n_ref: float, n_pred: float) -> dict:
    """Calcula Precision, Specificity y Error de Volumen a partir del JSON."""
    precision    = tp / (tp + fp)  if (tp + fp) > 0  else np.nan
    specificity  = tn / (tn + fp)  if (tn + fp) > 0  else np.nan
    vol_error    = (n_pred - n_ref) / n_ref * 100 if n_ref > 0 else np.nan
    return {
        "Precision":   precision,
        "Specificity": specificity,
        "Vol_Error_%": vol_error,
    }
 
 
# ------------------------------------------------------------------------------
# Procesado de cada caso
# ------------------------------------------------------------------------------
 
def process_case(case: dict, labels_dir: str, compute_spatial: bool = True) -> list:
    """
    Procesa un caso y devuelve una lista de dicts (uno por etiqueta).
 
    Parameters
    ----------
    case : dict
        Entrada de 'metric_per_case' del JSON de nnUNet.
    labels_dir : str
        Carpeta con los archivos de referencia (ground truth).
    compute_spatial : bool
        Si True, carga los NIfTIs y calcula HD95, HDmax, SSIM y Vol GT.
    """
    import nibabel as nib
 
    pred_file = case["prediction_file"]
    ref_file  = case["reference_file"]
    case_name = Path(pred_file).stem.replace(".nii", "")
 
    # Carga de NIfTIs 
    pred_data, ref_data, spacing, voxel_vol_ml = None, None, None, None
 
    if compute_spatial:
        pred_path = pred_file
        ref_path  = ref_file
 
        # Si la predicción no existe en la ruta original, buscar en labels_dir
        if not os.path.exists(pred_path):
            candidate = os.path.join(labels_dir, Path(pred_file).name)
            if os.path.exists(candidate):
                pred_path = candidate
 
        if os.path.exists(pred_path) and os.path.exists(ref_path):
            try:
                pred_nib   = nib.load(pred_path)
                ref_nib    = nib.load(ref_path)
                pred_data  = np.round(pred_nib.get_fdata()).astype(np.int16)
                ref_data   = np.round(ref_nib.get_fdata()).astype(np.int16)
                zooms      = ref_nib.header.get_zooms()[:3]          # mm
                spacing    = tuple(float(z) for z in zooms)
                voxel_vol_ml = float(np.prod(spacing)) / 1000.0      # mm3 -> mL
            except Exception as e:
                print(f"  [WARN] No se pudo cargar NIfTI para {case_name}: {e}")
        else:
            missing = pred_path if not os.path.exists(pred_path) else ref_path
            print(f"  [WARN] Archivo no encontrado: {missing}")
 
    # Métricas por etiqueta
    rows = []
    for label_str, m in case["metrics"].items():
        label = int(label_str)
 
        tp, fp, tn, fn = m["TP"], m["FP"], m["TN"], m["FN"]
        n_ref, n_pred  = m["n_ref"], m["n_pred"]
 
        row = {
            "Case":  case_name,
            "Label": label,
            "Dice":  round(m["Dice"], 6),
            "IoU":   round(m["IoU"],  6),
        }
 
        # Métricas del JSON
        derived = metrics_from_counts(tp, fp, tn, fn, n_ref, n_pred)
        row.update({k: round(v, 6) if not np.isnan(v) else np.nan
                    for k, v in derived.items()})
 
        # Métricas espaciales
        if voxel_vol_ml is not None and ref_data is not None:
            row["Vol_GT_mL"]   = round(n_ref  * voxel_vol_ml, 3)
            row["Vol_Pred_mL"] = round(n_pred * voxel_vol_ml, 3)
 
            pred_mask = (pred_data == label)
            ref_mask  = (ref_data  == label)
 
            hd95, hdmax = compute_hausdorff(pred_mask, ref_mask, spacing)
            row["HD95"]  = round(hd95,  3) if not np.isnan(hd95)  else np.nan
            row["HDmax"] = round(hdmax, 3) if not np.isnan(hdmax) else np.nan
            row["SSIM"]  = round(compute_ssim(pred_mask, ref_mask), 6)
        else:
            row["Vol_GT_mL"]   = np.nan
            row["Vol_Pred_mL"] = np.nan
            row["HD95"]        = np.nan
            row["HDmax"]       = np.nan
            row["SSIM"]        = np.nan
 
        rows.append(row)
 
    return rows
 
 
# ------------------------------------------------------------------------------
# Procesado completo de un JSON
# ------------------------------------------------------------------------------
 
def process_json(json_path: str, labels_dir: str,
                 model_tag: str, compute_spatial: bool = True) -> pd.DataFrame:
    """Lee el JSON de nnUNet y devuelve un DataFrame con todas las métricas."""
    try:
        from tqdm import tqdm
        use_tqdm = True
    except ImportError:
        use_tqdm = False
 
    with open(json_path, "r") as f:
        data = json.load(f)
 
    cases   = data["metric_per_case"]
    n_cases = len(cases)
    print(f"\n{'─'*60}")
    print(f"Modelo: {model_tag}  |  Casos: {n_cases}  |  JSON: {json_path}")
    print(f"{'─'*60}")
 
    all_rows = []
    iterator = tqdm(cases, desc=f"  Procesando {model_tag}", unit="caso") \
               if use_tqdm else cases
 
    for case in iterator:
        rows = process_case(case, labels_dir, compute_spatial=compute_spatial)
        all_rows.extend(rows)
 
    df = pd.DataFrame(all_rows)
    df.insert(0, "Model", model_tag)
    return df
 
 
# ------------------------------------------------------------------------------
# Tabla resumen (medias por modelo y etiqueta)
# ------------------------------------------------------------------------------
 
METRIC_COLS = ["Dice", "IoU", "Precision", "Specificity",
               "Vol_GT_mL", "Vol_Pred_mL", "Vol_Error_%", "HD95", "HDmax", "SSIM"]
 
 
def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula media y std de cada métrica, agrupando por Model y Label."""
    numeric = [c for c in METRIC_COLS if c in df.columns]
    grp = df.groupby(["Model", "Label"])[numeric]
    mean_df = grp.mean().round(4).add_suffix("_mean")
    std_df  = grp.std().round(4).add_suffix("_std")
    summary = pd.concat([mean_df, std_df], axis=1).reset_index()
 
    # Reordenar columnas: intercalar mean/std por métrica
    ordered = ["Model", "Label"]
    for col in numeric:
        ordered += [f"{col}_mean", f"{col}_std"]
    return summary[ordered]
 
 
# ------------------------------------------------------------------------------
# Exportación a Excel
# ------------------------------------------------------------------------------
 
def to_excel(df_per_case: pd.DataFrame, df_summary: pd.DataFrame,
             output_path: str) -> None:
    """Guarda los resultados en un archivo Excel con dos hojas."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import (Alignment, Font, PatternFill,
                                     numbers as xl_numbers)
        from openpyxl.utils import get_column_letter
 
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_per_case.to_excel(writer, sheet_name="Por_Caso",  index=False)
            df_summary.to_excel( writer, sheet_name="Resumen",   index=False)
 
            for sheet_name in ["Por_Caso", "Resumen"]:
                ws = writer.sheets[sheet_name]
                hdr_fill = PatternFill("solid", start_color="1F4E79")
                hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
 
                for cell in ws[1]:
                    cell.fill      = hdr_fill
                    cell.font      = hdr_font
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="center", wrap_text=True)
 
                # Auto-ancho
                for col_idx, col in enumerate(ws.columns, 1):
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value is not None),
                        default=8
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = \
                        min(max_len + 2, 22)
 
                # Zebra striping
                alt_fill = PatternFill("solid", start_color="EBF3FB")
                for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if i % 2 == 0:
                        for cell in row:
                            cell.fill = alt_fill
 
                ws.freeze_panes = "A2"
 
        print(f"\n Excel guardado: {output_path}")
 
    except ImportError:
        csv_path = output_path.replace(".xlsx", "_per_case.csv")
        sum_path = output_path.replace(".xlsx", "_summary.csv")
        df_per_case.to_csv(csv_path, index=False)
        df_summary.to_csv(sum_path,  index=False)
        print(f"\n CSVs guardados:\n  {csv_path}\n  {sum_path}")
 
 
# ------------------------------------------------------------------------------
# CLI
# ------------------------------------------------------------------------------
 
def parse_args():
    p = argparse.ArgumentParser(
        description="Calcula métricas de segmentación nnUNet (2D y/o 3D).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--json_3d",    type=str,  default=None,
                   help="Ruta al summary.json del modelo 3D.")
    p.add_argument("--json_2d",    type=str,  default=None,
                   help="Ruta al summary.json del modelo 2D.")
    p.add_argument("--labels_dir", type=str,
                   default="/home/ubuntu/nnUNet_raw/Dataset002_Kidney/labelsTs",
                   help="Carpeta con los labels GT (.nii.gz).")
    p.add_argument("--output",     type=str,  default="results_metrics.xlsx",
                   help="Archivo de salida (.xlsx o .csv).")
    p.add_argument("--no_spatial", action="store_true",
                   help="Omitir HD95/HDmax/SSIM (no carga NIfTIs, más rápido).")
    p.add_argument("--label_names", type=str, default=None,
                   help='JSON con mapeo de etiquetas, ej: \'{"1":"Kidney","2":"Tumor"}\'')
    return p.parse_args()
 
 
def main():
    args = parse_args()
 
    if args.json_3d is None and args.json_2d is None:
        print("[ERROR] Debes proporcionar al menos --json_3d o --json_2d.")
        return
 
    compute_spatial = not args.no_spatial
 
    # Mapeo de etiquetas (opcional)
    label_names = {}
    if args.label_names:
        label_names = json.loads(args.label_names)
 
    dfs = []
    if args.json_3d:
        dfs.append(process_json(args.json_3d, args.labels_dir,
                                model_tag="3D", compute_spatial=compute_spatial))
    if args.json_2d:
        dfs.append(process_json(args.json_2d, args.labels_dir,
                                model_tag="2D", compute_spatial=compute_spatial))
 
    df_all = pd.concat(dfs, ignore_index=True)
 
    # Añadir nombre de etiqueta si se proporcionó
    if label_names:
        df_all["Label_name"] = df_all["Label"].astype(str).map(label_names)
        df_all.insert(df_all.columns.get_loc("Label") + 1,
                      "Label_name", df_all.pop("Label_name"))
 
    df_summary = build_summary(df_all)
 
    # Impresión en consola 
    print("\n" + "═"*70)
    print("  RESUMEN DE MÉTRICAS (media ± std)")
    print("═"*70)
 
    summary_cols = ["Model", "Label"] + \
                   [c for c in df_summary.columns if c.endswith("_mean")]
    print(df_summary[summary_cols].to_string(index=False))
    print()
 
    # Comparativa 3D vs 2D 
    if args.json_3d and args.json_2d:
        print("═"*70)
        print("  DIFERENCIA 3D − 2D (media, por etiqueta)")
        print("═"*70)
        mean_cols = [c for c in df_summary.columns if c.endswith("_mean")]
        piv = df_summary.pivot_table(index="Label", columns="Model",
                                     values=mean_cols)
        for col in mean_cols:
            try:
                piv[("diff", col)] = piv[(col, "3D")] - piv[(col, "2D")]
            except KeyError:
                pass
        diff = piv.loc[:, "diff"].rename(
            columns={c: c.replace("_mean", "") for c in mean_cols}
        )
        print(diff.round(4).to_string())
        print()
 
    to_excel(df_all, df_summary, args.output)
 
 
if __name__ == "__main__":
    main()

